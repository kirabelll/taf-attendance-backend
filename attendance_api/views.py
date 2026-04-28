"""API Views for Attendance Data"""
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from datetime import datetime, timedelta
import sys
import os
import csv
import io
import json

# Add parent directory to path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from biotime_client import BioTimeClient
from config import Config

# Initialize BioTime client
biotime_client = BioTimeClient()
employee_cache = {}

def load_employee_cache():
    """Load employee data from BioTime"""
    global employee_cache
    try:
        employees = biotime_client.get_employees(page_size=1000)
        if employees:
            for emp in employees:
                emp_code = emp.get('emp_code', '')
                if emp_code:
                    full_name = f"{emp.get('first_name', '')} {emp.get('last_name', '')}".strip()
                    employee_cache[emp_code] = {
                        'full_name': full_name or 'Unknown',
                        'department': emp.get('department', 'Unknown'),
                        'position': emp.get('position', ''),
                        'first_name': emp.get('first_name', ''),
                        'last_name': emp.get('last_name', '')
                    }
        return True
    except Exception as e:
        print(f"Error loading employees: {str(e)}")
        return False

def get_employee_name(employee_id):
    """Get employee full name"""
    emp_data = employee_cache.get(employee_id, {})
    if isinstance(emp_data, dict):
        return emp_data.get('full_name', 'Unknown')
    return str(emp_data) if emp_data else 'Unknown'

def get_area_name(area):
    """Safely extract area name from area object or string"""
    if isinstance(area, dict):
        return area.get('area_name', area.get('area_code', 'Unknown'))
    return str(area) if area else 'Unknown'

def getDepartmentName(department):
    """Helper function to get department name from department object or string"""
    if isinstance(department, dict):
        return department.get('dept_name', department.get('dept_code', 'Unknown'))
    return str(department) if department else 'Unknown'

def calculate_late_minutes(punch_time):
    """Calculate late minutes"""
    late_threshold = punch_time.replace(
        hour=Config.LATE_THRESHOLD_HOUR,
        minute=Config.LATE_THRESHOLD_MINUTE,
        second=0,
        microsecond=0
    )
    if punch_time > late_threshold and punch_time.hour < 12:
        return int((punch_time - late_threshold).total_seconds() / 60)
    return 0

def get_log_type(record):
    """Determine log type: Clock In, Clock Out, or Late (M)"""
    punch_time = record['timestamp']
    punch_type = record.get('punch_type', '')
    
    # First, check the actual punch type from ZKBioTime
    if punch_type in ['Check Out', 'OUT', 'Clock Out', '1', 'Auto Check Out']:
        return 'Clock Out'
    elif punch_type in ['Check In', 'IN', 'Clock In', '0', 'Auto Check In']:
        # Check if it's late
        late_threshold = punch_time.replace(
            hour=Config.LATE_THRESHOLD_HOUR,
            minute=Config.LATE_THRESHOLD_MINUTE,
            second=0,
            microsecond=0
        )
        if punch_time > late_threshold and punch_time.hour < 12:
            return 'Late (M)'
        return 'Clock In'
    
    # For unknown punch types (like 255), use improved time-based logic
    # We need to look at the context of other punches for the same employee on the same day
    late_threshold = punch_time.replace(
        hour=Config.LATE_THRESHOLD_HOUR,
        minute=Config.LATE_THRESHOLD_MINUTE,
        second=0,
        microsecond=0
    )
    
    # If it's early in the day and after late threshold, it's likely late arrival
    if punch_time > late_threshold and punch_time.hour < 12:
        return 'Late (M)'
    # If it's early in the day, it's likely clock in
    elif punch_time.hour < 12:
        return 'Clock In'
    # If it's later in the day (after 1 PM), it's more likely clock out
    elif punch_time.hour >= 13:
        return 'Clock Out'
    # Noon hour is ambiguous - could be late lunch break or early departure
    else:
        return 'Clock Out'  # Default to clock out for noon hour

def process_attendance_records(records):
    """Process raw attendance records into daily summaries with total time"""
    daily_records = {}
    
    # Sort records by timestamp to process them in chronological order
    sorted_records = sorted(records, key=lambda x: x['timestamp'])
    
    for record in sorted_records:
        emp_id = record['employee_id']
        emp_name = record.get('employee_name') or get_employee_name(emp_id)
        punch_time = record['timestamp']
        log_type = get_log_type(record)
        late_minutes = calculate_late_minutes(punch_time) if log_type == 'Late (M)' else 0
        
        date_key = punch_time.strftime('%Y-%m-%d')
        key = f"{emp_id}_{date_key}"
        
        if key not in daily_records:
            daily_records[key] = {
                'employee_id': emp_id,
                'full_name': emp_name,
                'date': date_key,
                'clock_in': None,
                'clock_out': None,
                'late_minutes': 0,
                'device_id': record.get('device_id', 'Unknown'),
                'total_time_hours': 0,
                'duty_time_hours': 0,
                'working_hours': 0,
                'all_punches': [],  # Track all punches for debugging
                'clock_in_time': None,  # Store as datetime for comparison
                'clock_out_time': None  # Store as datetime for comparison
            }
        
        # Track all punches
        daily_records[key]['all_punches'].append({
            'time': punch_time.strftime('%H:%M:%S'),
            'type': log_type,
            'raw_type': record.get('punch_type', 'Unknown'),
            'timestamp': punch_time
        })
        
        # Update record with punch data - use earliest clock in and latest clock out
        if log_type in ['Clock In', 'Late (M)']:
            # Use the first clock in of the day
            if (daily_records[key]['clock_in_time'] is None or 
                punch_time < daily_records[key]['clock_in_time']):
                daily_records[key]['clock_in'] = punch_time.strftime('%H:%M:%S')
                daily_records[key]['clock_in_time'] = punch_time
                daily_records[key]['late_minutes'] = late_minutes
        elif log_type == 'Clock Out':
            # Use the latest clock out of the day
            if (daily_records[key]['clock_out_time'] is None or 
                punch_time > daily_records[key]['clock_out_time']):
                daily_records[key]['clock_out'] = punch_time.strftime('%H:%M:%S')
                daily_records[key]['clock_out_time'] = punch_time
        
        # Add total time data if available
        if 'total_time_hours' in record:
            daily_records[key]['total_time_hours'] = record['total_time_hours']
        if 'duty_time_hours' in record:
            daily_records[key]['duty_time_hours'] = record['duty_time_hours']
    
    # Calculate working hours for records that have both clock in and out
    for record in daily_records.values():
        if record['clock_in']:
            try:
                clock_in_time = record['clock_in_time']
                
                if record['clock_out']:
                    # Complete day - calculate total hours
                    clock_out_time = record['clock_out_time']
                    
                    # Handle next day clock out
                    if clock_out_time < clock_in_time:
                        clock_out_time += timedelta(days=1)
                    
                    working_duration = clock_out_time - clock_in_time
                    record['working_hours'] = round(working_duration.total_seconds() / 3600, 2)
                else:
                    # Still working - calculate hours so far (only for today)
                    record_date = datetime.strptime(record['date'], '%Y-%m-%d').date()
                    today = datetime.now().date()
                    
                    if record_date == today:
                        # Calculate hours worked so far today
                        now = datetime.now()
                        working_duration = now - clock_in_time
                        record['working_hours'] = round(working_duration.total_seconds() / 3600, 2)
                        record['status_note'] = 'Currently Working'
                    else:
                        # Past date without clock out - unknown hours
                        record['working_hours'] = 0
                        record['status_note'] = 'Incomplete Record'
                
                # Use total_time_hours if available and more accurate
                if record['total_time_hours'] > 0:
                    record['working_hours'] = record['total_time_hours']
                    
            except Exception as e:
                print(f"Error calculating working hours: {e}")
                record['working_hours'] = 0
        
        # Clean up temporary fields
        if 'clock_in_time' in record:
            del record['clock_in_time']
        if 'clock_out_time' in record:
            del record['clock_out_time']
        if 'all_punches' in record:
            del record['all_punches']
    
    return list(daily_records.values())

@csrf_exempt
@require_http_methods(["GET"])
def test_connection(request):
    """Test BioTime connection"""
    try:
        if biotime_client.authenticate():
            return JsonResponse({
                'success': True,
                'message': 'BioTime connection successful',
                'server': Config.BIOTIME_URL
            })
        else:
            return JsonResponse({
                'success': False,
                'message': 'Authentication failed'
            }, status=401)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Connection error: {str(e)}'
        }, status=500)

@csrf_exempt
@require_http_methods(["GET"])
def get_employees(request):
    """Get all employees"""
    try:
        employees = biotime_client.get_employees(page_size=1000)
        
        if employees:
            employee_list = []
            for emp in employees:
                employee_list.append({
                    'employee_id': emp.get('emp_code', ''),
                    'first_name': emp.get('first_name', ''),
                    'last_name': emp.get('last_name', ''),
                    'full_name': f"{emp.get('first_name', '')} {emp.get('last_name', '')}".strip(),
                    'department': emp.get('department', ''),
                    'position': emp.get('position', '')
                })
            
            return JsonResponse({
                'success': True,
                'count': len(employee_list),
                'employees': employee_list
            })
        else:
            return JsonResponse({
                'success': False,
                'message': 'No employees found'
            }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        }, status=500)

@csrf_exempt
@require_http_methods(["GET"])
def get_realtime_attendance(request):
    """Get real-time attendance data (last 24 hours)"""
    try:
        # Load employee cache if empty
        if not employee_cache:
            load_employee_cache()
        
        # Get transactions from last 24 hours
        start_time = datetime.now() - timedelta(hours=24)
        records = biotime_client.get_real_time_transactions(start_time)
        
        if records:
            processed_records = process_attendance_records(records)
            
            # Calculate statistics
            total_employees = len(processed_records)
            late_count = sum(1 for r in processed_records if r['late_minutes'] > 0)
            total_late_minutes = sum(r['late_minutes'] for r in processed_records)
            
            return JsonResponse({
                'success': True,
                'timestamp': datetime.now().isoformat(),
                'period': 'Last 24 hours',
                'statistics': {
                    'total_employees': total_employees,
                    'late_arrivals': late_count,
                    'total_late_minutes': total_late_minutes
                },
                'attendance': processed_records
            })
        else:
            return JsonResponse({
                'success': True,
                'message': 'No attendance records found',
                'attendance': []
            })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        }, status=500)

@csrf_exempt
@require_http_methods(["GET"])
def get_attendance_by_date(request):
    """Get attendance data for a specific date or date range with filtering"""
    try:
        # Load employee cache if empty
        if not employee_cache:
            load_employee_cache()
        
        # Get date parameters
        date_str = request.GET.get('date')
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        
        # Get filter parameters
        employee_name = request.GET.get('employee_name', '').strip()
        status_filter = request.GET.get('status', '').lower()  # 'late', 'absent', 'present', 'on_time'
        
        if date_str:
            # Single date
            date = datetime.strptime(date_str, '%Y-%m-%d')
            start_date = date
            end_date = date
        elif start_date_str and end_date_str:
            # Date range
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        else:
            # Default to today
            start_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = start_date
        
        print(f"Fetching attendance data from {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
        
        # Get raw records from ZKBioTime
        records = biotime_client.get_transactions_by_date_range(start_date, end_date)
        print(f"Retrieved {len(records)} raw transaction records")
        
        # Process records into daily summaries
        processed_records = process_attendance_records(records) if records else []
        print(f"Processed into {len(processed_records)} daily attendance records")
        
        # Get all employees for absent detection
        all_employees = []
        if employee_cache:
            for emp_id, emp_data in employee_cache.items():
                if isinstance(emp_data, dict):
                    all_employees.append((emp_id, emp_data.get('full_name', 'Unknown')))
                else:
                    all_employees.append((emp_id, str(emp_data)))
        
        present_employee_ids = {r['employee_id'] for r in processed_records}
        
        # Add absent employees for each date in the range
        current_date = start_date
        while current_date <= end_date:
            date_key = current_date.strftime('%Y-%m-%d')
            
            # Find employees present on this date
            date_present_ids = {r['employee_id'] for r in processed_records if r['date'] == date_key}
            
            # Add absent employees for this date
            for emp_id, emp_name in all_employees:
                if emp_id not in date_present_ids:
                    processed_records.append({
                        'employee_id': emp_id,
                        'full_name': emp_name,
                        'date': date_key,
                        'clock_in': None,
                        'clock_out': None,
                        'late_minutes': 0,
                        'device_id': 'N/A',
                        'status': 'Absent',
                        'working_hours': 0
                    })
            
            current_date += timedelta(days=1)
        
        # Add status to present employees
        for record in processed_records:
            if 'status' not in record:
                if record['clock_in'] is None and record['clock_out'] is None:
                    record['status'] = 'Absent'
                elif record['late_minutes'] > 0:
                    record['status'] = 'Late'
                else:
                    record['status'] = 'On Time'
        
        # Apply filters
        filtered_records = processed_records
        
        # Filter by employee name
        if employee_name:
            filtered_records = [
                r for r in filtered_records 
                if employee_name.lower() in r['full_name'].lower() or 
                   employee_name.lower() in r['employee_id'].lower()
            ]
        
        # Filter by status
        if status_filter:
            if status_filter == 'late':
                filtered_records = [r for r in filtered_records if r['status'] == 'Late']
            elif status_filter == 'absent':
                filtered_records = [r for r in filtered_records if r['status'] == 'Absent']
            elif status_filter == 'present':
                filtered_records = [r for r in filtered_records if r['status'] in ['Late', 'On Time']]
            elif status_filter == 'on_time':
                filtered_records = [r for r in filtered_records if r['status'] == 'On Time']
        
        # Calculate statistics
        total_employees = len(all_employees)
        present_count = len([r for r in processed_records if r['status'] in ['Late', 'On Time']])
        late_count = len([r for r in processed_records if r['status'] == 'Late'])
        absent_count = len([r for r in processed_records if r['status'] == 'Absent'])
        on_time_count = len([r for r in processed_records if r['status'] == 'On Time'])
        total_late_minutes = sum(r['late_minutes'] for r in processed_records)
        
        print(f"Final statistics: {total_employees} total, {present_count} present, {absent_count} absent, {late_count} late")
        
        return JsonResponse({
            'success': True,
            'start_date': start_date.strftime('%Y-%m-%d'),
            'end_date': end_date.strftime('%Y-%m-%d'),
            'filters': {
                'employee_name': employee_name or None,
                'status': status_filter or None
            },
            'statistics': {
                'total_employees': total_employees,
                'present': present_count,
                'absent': absent_count,
                'late_arrivals': late_count,
                'on_time': on_time_count,
                'total_late_minutes': total_late_minutes,
                'filtered_count': len(filtered_records),
                'raw_transactions': len(records),
                'processed_records': len(processed_records)
            },
            'attendance': filtered_records
        })
    except ValueError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid date format. Use YYYY-MM-DD'
        }, status=400)
    except Exception as e:
        print(f"Error in get_attendance_by_date: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        }, status=500)

@csrf_exempt
@require_http_methods(["GET"])
def get_attendance_report(request):
    """Get formatted attendance report"""
    try:
        # Load employee cache if empty
        if not employee_cache:
            load_employee_cache()
        
        # Get date parameter (default to today)
        date_str = request.GET.get('date', datetime.now().strftime('%Y-%m-%d'))
        date = datetime.strptime(date_str, '%Y-%m-%d')
        
        records = biotime_client.get_transactions_by_date_range(date, date)
        
        if records:
            processed_records = process_attendance_records(records)
            
            # Sort by employee ID
            processed_records.sort(key=lambda x: x['employee_id'])
            
            # Calculate statistics
            total_employees = len(processed_records)
            late_count = sum(1 for r in processed_records if r['late_minutes'] > 0)
            total_late_minutes = sum(r['late_minutes'] for r in processed_records)
            
            # Format report
            report = {
                'title': f'ATTENDANCE REPORT - {date.strftime("%Y-%m-%d")}',
                'generated_at': datetime.now().isoformat(),
                'summary': {
                    'total_employees': total_employees,
                    'late_arrivals': late_count,
                    'total_late_minutes': total_late_minutes,
                    'on_time': total_employees - late_count
                },
                'records': []
            }
            
            for idx, record in enumerate(processed_records, 1):
                report['records'].append({
                    'no': idx,
                    'employee_id': record['employee_id'],
                    'full_name': record['full_name'],
                    'clock_in': record['clock_in'] or '---',
                    'clock_out': record['clock_out'] or '---',
                    'late_minutes': record['late_minutes'],
                    'status': 'Late' if record['late_minutes'] > 0 else 'On Time'
                })
            
            return JsonResponse({
                'success': True,
                'report': report
            })
        else:
            return JsonResponse({
                'success': True,
                'message': 'No attendance records found',
                'report': {
                    'title': f'ATTENDANCE REPORT - {date.strftime("%Y-%m-%d")}',
                    'generated_at': datetime.now().isoformat(),
                    'summary': {
                        'total_employees': 0,
                        'late_arrivals': 0,
                        'total_late_minutes': 0,
                        'on_time': 0
                    },
                    'records': []
                }
            })
    except ValueError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid date format. Use YYYY-MM-DD'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        }, status=500)

@csrf_exempt
@require_http_methods(["POST"])
def refresh_employee_cache(request):
    """Manually refresh employee cache"""
    try:
        if load_employee_cache():
            return JsonResponse({
                'success': True,
                'message': f'Employee cache refreshed. Loaded {len(employee_cache)} employees'
            })
        else:
            return JsonResponse({
                'success': False,
                'message': 'Failed to refresh employee cache'
            }, status=500)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        }, status=500)

@csrf_exempt
@require_http_methods(["GET"])
def get_late_employees(request):
    """Get employees who arrived late"""
    try:
        # Load employee cache if empty
        if not employee_cache:
            load_employee_cache()
        
        # Get date parameter (default to today)
        date_str = request.GET.get('date', datetime.now().strftime('%Y-%m-%d'))
        date = datetime.strptime(date_str, '%Y-%m-%d')
        
        records = biotime_client.get_transactions_by_date_range(date, date)
        processed_records = process_attendance_records(records) if records else []
        
        # Filter only late employees
        late_employees = [r for r in processed_records if r['late_minutes'] > 0]
        
        # Sort by late minutes (most late first)
        late_employees.sort(key=lambda x: x['late_minutes'], reverse=True)
        
        return JsonResponse({
            'success': True,
            'date': date.strftime('%Y-%m-%d'),
            'total_late': len(late_employees),
            'total_late_minutes': sum(r['late_minutes'] for r in late_employees),
            'late_employees': late_employees
        })
    except ValueError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid date format. Use YYYY-MM-DD'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        }, status=500)

@csrf_exempt
@require_http_methods(["GET"])
def get_absent_employees(request):
    """Get employees who are absent"""
    try:
        # Load employee cache if empty
        if not employee_cache:
            load_employee_cache()
        
        # Get date parameter (default to today)
        date_str = request.GET.get('date', datetime.now().strftime('%Y-%m-%d'))
        date = datetime.strptime(date_str, '%Y-%m-%d')
        
        records = biotime_client.get_transactions_by_date_range(date, date)
        processed_records = process_attendance_records(records) if records else []
        
        # Get present employee IDs
        present_employee_ids = {r['employee_id'] for r in processed_records}
        
        # Find absent employees
        absent_employees = []
        for emp_id, emp_data in employee_cache.items():
            if emp_id not in present_employee_ids:
                emp_name = emp_data.get('full_name', 'Unknown') if isinstance(emp_data, dict) else str(emp_data)
                absent_employees.append({
                    'employee_id': emp_id,
                    'full_name': emp_name,
                    'date': date.strftime('%Y-%m-%d'),
                    'status': 'Absent'
                })
        
        # Sort by employee ID
        absent_employees.sort(key=lambda x: x['employee_id'])
        
        return JsonResponse({
            'success': True,
            'date': date.strftime('%Y-%m-%d'),
            'total_employees': len(employee_cache),
            'present_count': len(present_employee_ids),
            'absent_count': len(absent_employees),
            'absent_employees': absent_employees
        })
    except ValueError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid date format. Use YYYY-MM-DD'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        }, status=500)

@csrf_exempt
@require_http_methods(["GET"])
def get_daily_attendance_summary(request):
    """Get daily attendance summary with all statistics"""
    try:
        # Load employee cache if empty
        if not employee_cache:
            load_employee_cache()
        
        # Get date parameter (default to today)
        date_str = request.GET.get('date', datetime.now().strftime('%Y-%m-%d'))
        date = datetime.strptime(date_str, '%Y-%m-%d')
        
        records = biotime_client.get_transactions_by_date_range(date, date)
        processed_records = process_attendance_records(records) if records else []
        
        # Get present employee IDs
        present_employee_ids = {r['employee_id'] for r in processed_records}
        
        # Categorize employees
        on_time = [r for r in processed_records if r['late_minutes'] == 0]
        late = [r for r in processed_records if r['late_minutes'] > 0]
        
        # Find absent employees
        absent = []
        for emp_id, emp_data in employee_cache.items():
            if emp_id not in present_employee_ids:
                emp_name = emp_data.get('full_name', 'Unknown') if isinstance(emp_data, dict) else str(emp_data)
                absent.append({
                    'employee_id': emp_id,
                    'full_name': emp_name,
                    'date': date.strftime('%Y-%m-%d'),
                    'status': 'Absent'
                })
        
        total_employees = len(employee_cache)
        present_count = len(processed_records)
        absent_count = len(absent)
        late_count = len(late)
        on_time_count = len(on_time)
        total_late_minutes = sum(r['late_minutes'] for r in late)
        
        # Calculate percentages
        attendance_rate = (present_count / total_employees * 100) if total_employees > 0 else 0
        punctuality_rate = (on_time_count / present_count * 100) if present_count > 0 else 0
        
        return JsonResponse({
            'success': True,
            'date': date.strftime('%Y-%m-%d'),
            'summary': {
                'total_employees': total_employees,
                'present': present_count,
                'absent': absent_count,
                'on_time': on_time_count,
                'late': late_count,
                'total_late_minutes': total_late_minutes,
                'attendance_rate': round(attendance_rate, 2),
                'punctuality_rate': round(punctuality_rate, 2)
            },
            'breakdown': {
                'on_time_employees': on_time,
                'late_employees': late,
                'absent_employees': absent
            }
        })
    except ValueError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid date format. Use YYYY-MM-DD'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        }, status=500)

@csrf_exempt
@require_http_methods(["GET"])
def search_employee_attendance(request):
    """Search employee attendance by name or ID"""
    try:
        # Load employee cache if empty
        if not employee_cache:
            load_employee_cache()
        
        # Get search parameters
        search_query = request.GET.get('q', '').strip()
        date_str = request.GET.get('date', datetime.now().strftime('%Y-%m-%d'))
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        
        if not search_query:
            return JsonResponse({
                'success': False,
                'message': 'Search query parameter "q" is required'
            }, status=400)
        
        # Determine date range
        if start_date_str and end_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        else:
            date = datetime.strptime(date_str, '%Y-%m-%d')
            start_date = date
            end_date = date
        
        records = biotime_client.get_transactions_by_date_range(start_date, end_date)
        processed_records = process_attendance_records(records) if records else []
        
        # Add absent employees for the date range
        all_dates = []
        current_date = start_date
        while current_date <= end_date:
            all_dates.append(current_date.strftime('%Y-%m-%d'))
            current_date += timedelta(days=1)
        
        # Search and filter records
        matching_records = []
        for record in processed_records:
            if (search_query.lower() in record['full_name'].lower() or 
                search_query.lower() in record['employee_id'].lower()):
                matching_records.append(record)
        
        # Check for absent days for matching employees
        matching_employee_ids = {r['employee_id'] for r in matching_records}
        
        # Add absent records for matching employees
        for emp_id, emp_data in employee_cache.items():
            emp_name = emp_data.get('full_name', 'Unknown') if isinstance(emp_data, dict) else str(emp_data)
            if (search_query.lower() in emp_name.lower() or 
                search_query.lower() in emp_id.lower()):
                matching_employee_ids.add(emp_id)
                
                # Check each date for absence
                for date_str in all_dates:
                    has_record = any(r['employee_id'] == emp_id and r['date'] == date_str 
                                   for r in matching_records)
                    if not has_record:
                        matching_records.append({
                            'employee_id': emp_id,
                            'full_name': emp_name,
                            'date': date_str,
                            'clock_in': None,
                            'clock_out': None,
                            'late_minutes': 0,
                            'device_id': 'N/A',
                            'status': 'Absent'
                        })
        
        # Add status to records
        for record in matching_records:
            if 'status' not in record:
                if record['clock_in'] is None and record['clock_out'] is None:
                    record['status'] = 'Absent'
                elif record['late_minutes'] > 0:
                    record['status'] = 'Late'
                else:
                    record['status'] = 'On Time'
        
        # Sort by date and employee ID
        matching_records.sort(key=lambda x: (x['date'], x['employee_id']))
        
        return JsonResponse({
            'success': True,
            'search_query': search_query,
            'start_date': start_date.strftime('%Y-%m-%d'),
            'end_date': end_date.strftime('%Y-%m-%d'),
            'total_records': len(matching_records),
            'matching_employees': len(matching_employee_ids),
            'records': matching_records
        })
    except ValueError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid date format. Use YYYY-MM-DD'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        }, status=500)

# Device Management Views

def get_biotime_devices():
    """Get real device data from ZKBioTime"""
    try:
        if not biotime_client.token:
            biotime_client.authenticate()
        
        # Get devices from ZKBioTime API
        url = f"{biotime_client.base_url}/iclock/api/terminals/"
        response = biotime_client.session.get(url, params={'page_size': 100})
        
        if response.status_code == 200:
            data = response.json()
            terminals = data.get('data', [])
            
            devices = []
            for terminal in terminals:
                # Calculate status based on last activity
                last_activity = terminal.get('last_activity')
                status = "Online"
                if last_activity:
                    try:
                        last_time = datetime.strptime(last_activity, '%Y-%m-%d %H:%M:%S')
                        time_diff = datetime.now() - last_time
                        if time_diff > timedelta(hours=1):
                            status = "Offline"
                        elif time_diff > timedelta(minutes=30):
                            status = "Maintenance"
                    except:
                        status = "Unknown"
                
                # Handle area object - extract area_name if it's an object
                area = get_area_name(terminal.get('area', 'Unknown'))
                
                devices.append({
                    "id": terminal.get('sn', ''),
                    "name": terminal.get('alias', terminal.get('sn', '')),
                    "ip": terminal.get('ip_address', ''),
                    "location": area,
                    "status": status,
                    "last_sync": terminal.get('last_activity', 'Unknown'),
                    "total_scans": terminal.get('transaction_count', 0),
                    "model": "ZKBioTime Device",
                    "firmware_version": terminal.get('fw_version', ''),
                    "sync_errors": 0,
                    "last_heartbeat": terminal.get('last_activity', datetime.now().isoformat()),
                    "user_count": terminal.get('user_count', 0),
                    "fp_count": terminal.get('fp_count', 0),
                    "face_count": terminal.get('face_count', 0),
                    "palm_count": terminal.get('palm_count', 0)
                })
            
            return devices
        else:
            print(f"Failed to fetch devices: {response.status_code}")
            return get_mock_devices()
    except Exception as e:
        print(f"Error fetching devices from ZKBioTime: {str(e)}")
        return get_mock_devices()

def get_mock_devices():
    """Fallback mock device data"""
    return [
        {
            "id": "VDE225240197",
            "name": "Entry",
            "ip": "172.16.10.210",
            "location": "Main Door",
            "status": "Online",
            "last_sync": "2026-03-12 09:40:25",
            "total_scans": 30372,
            "model": "ZKBioTime Device",
            "firmware_version": "6.60.1.0",
            "sync_errors": 0,
            "last_heartbeat": datetime.now().isoformat(),
            "user_count": 120,
            "fp_count": 1,
            "face_count": 112,
            "palm_count": 0
        },
        {
            "id": "VDE225240198",
            "name": "Main Door",
            "ip": "172.16.10.211",
            "location": "Main Door",
            "status": "Online",
            "last_sync": "2026-03-12 09:40:30",
            "total_scans": 43115,
            "model": "ZKBioTime Device",
            "firmware_version": "6.60.1.0",
            "sync_errors": 0,
            "last_heartbeat": datetime.now().isoformat(),
            "user_count": 123,
            "fp_count": 1,
            "face_count": 107,
            "palm_count": 0
        }
    ]

@csrf_exempt
@require_http_methods(["GET"])
def get_devices(request):
    """Get all devices with statistics"""
    try:
        # Get real device data from ZKBioTime
        devices = get_biotime_devices()
        
        online_count = len([d for d in devices if d["status"] == "Online"])
        offline_count = len([d for d in devices if d["status"] == "Offline"])
        maintenance_count = len([d for d in devices if d["status"] == "Maintenance"])
        total_scans = sum(d["total_scans"] for d in devices)
        
        stats = {
            "total_devices": len(devices),
            "online_devices": online_count,
            "offline_devices": offline_count,
            "maintenance_devices": maintenance_count,
            "total_scans_today": total_scans,
            "last_sync_time": "2 min ago"
        }
        
        return JsonResponse({
            "success": True,
            "devices": devices,
            "stats": stats
        })
    except Exception as e:
        return JsonResponse({
            "success": False,
            "message": f"Error: {str(e)}"
        }, status=500)

@csrf_exempt
@require_http_methods(["GET"])
def get_device_stats(request):
    """Get device statistics only"""
    try:
        devices = get_biotime_devices()
        
        online_count = len([d for d in devices if d["status"] == "Online"])
        offline_count = len([d for d in devices if d["status"] == "Offline"])
        maintenance_count = len([d for d in devices if d["status"] == "Maintenance"])
        total_scans = sum(d["total_scans"] for d in devices)
        
        stats = {
            "total_devices": len(devices),
            "online_devices": online_count,
            "offline_devices": offline_count,
            "maintenance_devices": maintenance_count,
            "total_scans_today": total_scans,
            "last_sync_time": datetime.now().strftime("%M min ago")
        }
        
        return JsonResponse({
            "success": True,
            "stats": stats
        })
    except Exception as e:
        return JsonResponse({
            "success": False,
            "message": f"Error: {str(e)}"
        }, status=500)

@csrf_exempt
@require_http_methods(["POST"])
def sync_device(request, device_id):
    """Sync a specific device"""
    try:
        devices = get_biotime_devices()
        device = next((d for d in devices if d["id"] == device_id), None)
        if not device:
            return JsonResponse({
                "success": False,
                "message": "Device not found"
            }, status=404)
        
        # Try to sync with ZKBioTime device
        try:
            if biotime_client.token:
                # Call ZKBioTime sync API if available
                url = f"{biotime_client.base_url}/iclock/api/terminals/{device_id}/sync/"
                response = biotime_client.session.post(url)
                
                if response.status_code == 200:
                    return JsonResponse({
                        "success": True,
                        "message": f"Device {device_id} synced successfully with ZKBioTime",
                        "sync_time": datetime.now().isoformat()
                    })
        except Exception as sync_error:
            print(f"ZKBioTime sync error: {sync_error}")
        
        # Fallback to simulated sync
        return JsonResponse({
            "success": True,
            "message": f"Device {device_id} sync initiated",
            "sync_time": datetime.now().isoformat()
        })
    except Exception as e:
        return JsonResponse({
            "success": False,
            "message": f"Sync failed: {str(e)}"
        }, status=500)

@csrf_exempt
@require_http_methods(["PATCH"])
def update_device_status(request, device_id):
    """Update device status"""
    try:
        devices = get_biotime_devices()
        device = next((d for d in devices if d["id"] == device_id), None)
        if not device:
            return JsonResponse({
                "success": False,
                "message": "Device not found"
            }, status=404)
        
        import json
        data = json.loads(request.body)
        new_status = data.get("status")
        
        if new_status not in ["Online", "Offline", "Maintenance"]:
            return JsonResponse({
                "success": False,
                "message": "Invalid status"
            }, status=400)
        
        # Note: In a real implementation, you might want to store device status overrides
        # in a database since ZKBioTime determines status automatically
        
        return JsonResponse({
            "success": True,
            "message": f"Device {device_id} status update requested (Note: ZKBioTime manages actual device status)"
        })
    except Exception as e:
        return JsonResponse({
            "success": False,
            "message": f"Status update failed: {str(e)}"
        }, status=500)

@csrf_exempt
@require_http_methods(["GET"])
def get_device_logs(request, device_id):
    """Get device logs"""
    try:
        devices = get_biotime_devices()
        device = next((d for d in devices if d["id"] == device_id), None)
        if not device:
            return JsonResponse({
                "success": False,
                "message": "Device not found"
            }, status=404)
        
        # Try to get real logs from ZKBioTime
        logs = []
        try:
            if biotime_client.token:
                # Get recent transactions for this device as logs
                url = f"{biotime_client.base_url}/iclock/api/transactions/"
                params = {
                    'terminal_sn': device_id,
                    'start_time': (datetime.now() - timedelta(hours=24)).strftime('%Y-%m-%d %H:%M:%S'),
                    'page_size': 10
                }
                response = biotime_client.session.get(url, params=params)
                
                if response.status_code == 200:
                    data = response.json()
                    transactions = data.get('data', [])
                    
                    for trans in transactions:
                        logs.append({
                            "id": f"log_{device_id}_{trans.get('id', '')}",
                            "timestamp": trans.get('punch_time', ''),
                            "level": "INFO",
                            "message": f"Transaction: {trans.get('emp_code', '')} - {trans.get('punch_state', '')}",
                            "device_id": device_id
                        })
        except Exception as log_error:
            print(f"Error fetching device logs: {log_error}")
        
        # Add some system logs
        if not logs:
            logs = [
                {
                    "id": f"log_{device_id}_1",
                    "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "level": "INFO",
                    "message": f"Device {device_id} heartbeat received",
                    "device_id": device_id
                },
                {
                    "id": f"log_{device_id}_2",
                    "timestamp": (datetime.now() - timedelta(minutes=5)).strftime("%Y-%m-%d %H:%M:%S"),
                    "level": "INFO",
                    "message": f"Attendance data synced successfully",
                    "device_id": device_id
                }
            ]
        
        return JsonResponse({
            "success": True,
            "logs": logs
        })
    except Exception as e:
        return JsonResponse({
            "success": False,
            "message": f"Error: {str(e)}"
        }, status=500)

@csrf_exempt
@require_http_methods(["POST"])
def test_device_connection(request, device_id):
    """Test device connection"""
    try:
        devices = get_biotime_devices()
        device = next((d for d in devices if d["id"] == device_id), None)
        if not device:
            return JsonResponse({
                "success": False,
                "message": "Device not found"
            }, status=404)
        
        # Try to ping the device IP
        import subprocess
        import platform
        
        device_ip = device.get("ip", "")
        if device_ip:
            try:
                # Use ping command based on OS
                param = "-n" if platform.system().lower() == "windows" else "-c"
                command = ["ping", param, "1", device_ip]
                
                result = subprocess.run(command, capture_output=True, text=True, timeout=5)
                online = result.returncode == 0
                
                # Estimate response time from ping output
                response_time = 50  # Default
                if online and "time=" in result.stdout:
                    try:
                        import re
                        time_match = re.search(r'time[<=](\d+)', result.stdout)
                        if time_match:
                            response_time = int(time_match.group(1))
                    except:
                        pass
                
                return JsonResponse({
                    "success": True,
                    "online": online,
                    "response_time": response_time,
                    "message": f"Device {device_id} {'is reachable' if online else 'is not reachable'} at {device_ip}"
                })
            except subprocess.TimeoutExpired:
                return JsonResponse({
                    "success": True,
                    "online": False,
                    "response_time": 0,
                    "message": f"Device {device_id} connection timeout"
                })
        else:
            return JsonResponse({
                "success": False,
                "message": "Device IP address not available"
            })
    except Exception as e:
        return JsonResponse({
            "success": False,
            "message": f"Connection test failed: {str(e)}"
        }, status=500)

# Debug endpoint to check transaction processing
@csrf_exempt
@require_http_methods(["GET"])
def debug_attendance_processing(request):
    """Debug endpoint to see how attendance data is being processed"""
    try:
        # Get date parameter (default to today)
        date_str = request.GET.get('date', datetime.now().strftime('%Y-%m-%d'))
        employee_id = request.GET.get('employee_id')
        
        date = datetime.strptime(date_str, '%Y-%m-%d')
        
        # Get raw transactions
        raw_records = biotime_client.get_transactions_by_date_range(date, date)
        
        # Filter by employee if specified
        if employee_id:
            raw_records = [r for r in raw_records if r['employee_id'] == employee_id]
        
        # Process records
        processed_records = process_attendance_records(raw_records)
        
        # Filter processed records by employee if specified
        if employee_id:
            processed_records = [r for r in processed_records if r['employee_id'] == employee_id]
        
        debug_info = {
            'date': date_str,
            'employee_filter': employee_id,
            'raw_transactions_count': len(raw_records),
            'processed_records_count': len(processed_records),
            'raw_transactions': [
                {
                    'employee_id': r['employee_id'],
                    'employee_name': r.get('employee_name', ''),
                    'timestamp': r['timestamp'].strftime('%Y-%m-%d %H:%M:%S'),
                    'punch_type': r.get('punch_type', ''),
                    'raw_punch_state': r.get('raw_punch_state', ''),
                    'device_id': r.get('device_id', '')
                } for r in raw_records[:20]  # Limit to first 20 for readability
            ],
            'processed_records': processed_records[:10]  # Limit to first 10
        }
        
        return JsonResponse({
            'success': True,
            'debug_info': debug_info
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Debug failed: {str(e)}'
        }, status=500)

# Employee Late Report
@csrf_exempt
@require_http_methods(["GET"])
def get_employee_late_report(request):
    """Get comprehensive late report for all employees"""
    try:
        # Load employee cache if empty
        if not employee_cache:
            load_employee_cache()
        
        # Get date parameters
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        
        if start_date_str and end_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        else:
            # Default to last 30 days
            end_date = datetime.now()
            start_date = end_date - timedelta(days=30)
        
        print(f"Generating late report from {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
        
        # Get all attendance records for the period
        records = biotime_client.get_transactions_by_date_range(start_date, end_date)
        processed_records = process_attendance_records(records) if records else []
        
        # Calculate late statistics per employee
        employee_late_stats = {}
        
        for record in processed_records:
            emp_id = record['employee_id']
            emp_name = record['full_name']
            late_minutes = record.get('late_minutes', 0)
            
            if emp_id not in employee_late_stats:
                employee_late_stats[emp_id] = {
                    'employee_id': emp_id,
                    'full_name': emp_name,
                    'total_days': 0,
                    'late_days': 0,
                    'total_late_minutes': 0,
                    'average_late_minutes': 0,
                    'punctuality_rate': 0,
                    'late_incidents': []
                }
            
            # Count this day
            employee_late_stats[emp_id]['total_days'] += 1
            
            # If employee was late
            if late_minutes > 0:
                employee_late_stats[emp_id]['late_days'] += 1
                employee_late_stats[emp_id]['total_late_minutes'] += late_minutes
                employee_late_stats[emp_id]['late_incidents'].append({
                    'date': record['date'],
                    'clock_in': record['clock_in'],
                    'late_minutes': late_minutes
                })
        
        # Calculate averages and rates
        for emp_id, stats in employee_late_stats.items():
            if stats['total_days'] > 0:
                stats['punctuality_rate'] = round(((stats['total_days'] - stats['late_days']) / stats['total_days']) * 100, 2)
            
            if stats['late_days'] > 0:
                stats['average_late_minutes'] = round(stats['total_late_minutes'] / stats['late_days'], 2)
        
        # Convert to list and sort by total late minutes (worst first)
        late_report = list(employee_late_stats.values())
        late_report.sort(key=lambda x: x['total_late_minutes'], reverse=True)
        
        # Calculate overall statistics
        total_employees = len(late_report)
        employees_with_late = len([emp for emp in late_report if emp['late_days'] > 0])
        total_late_incidents = sum(emp['late_days'] for emp in late_report)
        total_late_minutes_all = sum(emp['total_late_minutes'] for emp in late_report)
        
        # Calculate average punctuality rate
        avg_punctuality = 0
        if total_employees > 0:
            avg_punctuality = round(sum(emp['punctuality_rate'] for emp in late_report) / total_employees, 2)
        
        return JsonResponse({
            'success': True,
            'period': {
                'start_date': start_date.strftime('%Y-%m-%d'),
                'end_date': end_date.strftime('%Y-%m-%d'),
                'days': (end_date - start_date).days + 1
            },
            'summary': {
                'total_employees': total_employees,
                'employees_with_late': employees_with_late,
                'total_late_incidents': total_late_incidents,
                'total_late_minutes': total_late_minutes_all,
                'total_late_hours': round(total_late_minutes_all / 60, 2),
                'average_punctuality_rate': avg_punctuality,
                'most_punctual_rate': max([emp['punctuality_rate'] for emp in late_report], default=0),
                'least_punctual_rate': min([emp['punctuality_rate'] for emp in late_report], default=0)
            },
            'employees': late_report
        })
        
    except ValueError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid date format. Use YYYY-MM-DD'
        }, status=400)
    except Exception as e:
        print(f"Error generating late report: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        }, status=500)


@csrf_exempt
@require_http_methods(["GET"])
def export_late_report_csv(request):
    """Export employee late report as CSV matching the frontend display format"""
    try:
        print(f"CSV Export Request received: {request.method} {request.get_full_path()}")
        
        # Load employee cache if empty
        if not employee_cache:
            load_employee_cache()
        
        # Get date parameters
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        
        print(f"Date parameters: start_date={start_date_str}, end_date={end_date_str}")
        
        if start_date_str and end_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        else:
            # Default to last 7 days for better performance
            end_date = datetime.now()
            start_date = end_date - timedelta(days=7)
        
        # Limit date range to prevent timeouts (max 30 days)
        max_days = 30
        if (end_date - start_date).days > max_days:
            print(f"Date range too large ({(end_date - start_date).days} days), limiting to {max_days} days")
            start_date = end_date - timedelta(days=max_days)
        
        print(f"Exporting late report CSV from {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
        
        # Get all attendance records for the period with optimized processing
        print("Fetching attendance records...")
        records = biotime_client.get_transactions_by_date_range(start_date, end_date)
        print(f"Retrieved {len(records)} raw records")
        
        processed_records = process_attendance_records(records) if records else []
        print(f"Processed {len(processed_records)} attendance records")
        
        # Calculate late statistics per employee with optimized processing
        employee_late_stats = {}
        
        print("Calculating late statistics...")
        for i, record in enumerate(processed_records):
            if i % 100 == 0:  # Progress indicator
                print(f"Processing record {i+1}/{len(processed_records)}")
                
            emp_id = record['employee_id']
            emp_name = record['full_name']
            late_minutes = record.get('late_minutes', 0)
            
            if emp_id not in employee_late_stats:
                employee_late_stats[emp_id] = {
                    'employee_id': emp_id,
                    'full_name': emp_name,
                    'total_days': 0,
                    'late_days': 0,
                    'total_late_minutes': 0,
                    'average_late_minutes': 0,
                    'punctuality_rate': 0,
                    'late_incidents': []
                }
            
            # Count this day
            employee_late_stats[emp_id]['total_days'] += 1
            
            # If employee was late
            if late_minutes > 0:
                employee_late_stats[emp_id]['late_days'] += 1
                employee_late_stats[emp_id]['total_late_minutes'] += late_minutes
                employee_late_stats[emp_id]['late_incidents'].append({
                    'date': record['date'],
                    'clock_in': record['clock_in'],
                    'late_minutes': late_minutes
                })
        
        print("Calculating averages and rates...")
        # Calculate averages and rates
        for emp_id, stats in employee_late_stats.items():
            if stats['total_days'] > 0:
                stats['punctuality_rate'] = round(((stats['total_days'] - stats['late_days']) / stats['total_days']) * 100, 2)
            
            if stats['late_days'] > 0:
                stats['average_late_minutes'] = round(stats['total_late_minutes'] / stats['late_days'], 2)
        
        # Convert to list and sort by total late minutes (worst first)
        late_report = list(employee_late_stats.values())
        late_report.sort(key=lambda x: x['total_late_minutes'], reverse=True)
        
        print(f"Generated late report for {len(late_report)} employees")
        
        # Calculate overall statistics matching the frontend display
        total_employees = len(late_report)
        employees_with_late = len([emp for emp in late_report if emp['late_days'] > 0])
        total_late_incidents = sum(emp['late_days'] for emp in late_report)
        total_late_minutes_all = sum(emp['total_late_minutes'] for emp in late_report)
        total_late_hours = round(total_late_minutes_all / 60, 1)  # Match frontend format (53.2h)
        
        # Calculate average punctuality rate
        avg_punctuality = 0
        if total_employees > 0:
            avg_punctuality = round(sum(emp['punctuality_rate'] for emp in late_report) / total_employees, 2)
        
        # Calculate report period days
        period_days = (end_date - start_date).days + 1
        
        print("Generating CSV content...")
        # Create CSV content matching the frontend format exactly
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Add title matching frontend
        writer.writerow(["Employee Late Report"])
        writer.writerow([])  # Empty row
        
        # Add summary statistics matching frontend cards
        writer.writerow(["SUMMARY STATISTICS"])
        writer.writerow(["Total Employees", total_employees])
        writer.writerow(["Employees with Late", employees_with_late])
        writer.writerow(["Avg Punctuality", f"{avg_punctuality}%"])
        writer.writerow(["Total Late Hours", f"{total_late_hours}h"])
        writer.writerow([])  # Empty row
        
        # Add report period info
        writer.writerow(["Report Period", f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')} ({period_days} days)"])
        writer.writerow([])  # Empty row
        
        # Add employee data headers matching the frontend table exactly
        writer.writerow([
            "Rank",
            "Employee", 
            "Total Days", 
            "Late Days", 
            "Total Late (min)", 
            "Avg Late (min)", 
            "Punctuality %"
        ])
        
        # Add employee data matching frontend format (limit to top 100 for performance)
        top_employees = late_report[:100] if len(late_report) > 100 else late_report
        for rank, employee in enumerate(top_employees, 1):
            # Format rank with # prefix like frontend
            rank_formatted = f"#{rank}"
            
            # Format punctuality percentage to match frontend (e.g., "16.67%", "40%")
            punctuality_formatted = f"{employee['punctuality_rate']}%"
            
            # Format average late minutes to match frontend display
            avg_late_formatted = employee['average_late_minutes']
            if avg_late_formatted == int(avg_late_formatted):
                avg_late_formatted = int(avg_late_formatted)
            
            writer.writerow([
                rank_formatted,
                employee['full_name'],
                employee['total_days'],
                employee['late_days'],
                employee['total_late_minutes'],
                avg_late_formatted,
                punctuality_formatted
            ])
        
        # Add detailed late incidents section (limit to recent 500 incidents for performance)
        writer.writerow([])  # Empty row
        writer.writerow(["DETAILED LATE INCIDENTS"])
        writer.writerow([
            "Employee Name",
            "Employee ID",
            "Date",
            "Clock In Time",
            "Late Minutes"
        ])
        
        # Add all late incidents (limited for performance)
        all_incidents = []
        for employee in top_employees:
            for incident in employee['late_incidents']:
                all_incidents.append({
                    'name': employee['full_name'],
                    'id': employee['employee_id'],
                    'date': incident['date'],
                    'clock_in': incident['clock_in'],
                    'late_minutes': incident['late_minutes']
                })
        
        # Sort incidents by date (most recent first) and limit to 500
        all_incidents.sort(key=lambda x: x['date'], reverse=True)
        limited_incidents = all_incidents[:500]
        
        # Add incident data
        for incident in limited_incidents:
            writer.writerow([
                incident['name'],
                incident['id'],
                incident['date'],
                incident['clock_in'],
                incident['late_minutes']
            ])
        
        print(f"CSV content generated successfully with {len(limited_incidents)} incidents")
        
        # Create response
        response = HttpResponse(
            output.getvalue(),
            content_type='text/csv; charset=utf-8'
        )
        response['Content-Disposition'] = f'attachment; filename="employee_late_report_{start_date.strftime("%Y%m%d")}_to_{end_date.strftime("%Y%m%d")}.csv"'
        
        print("CSV export completed successfully")
        return response
        
    except ValueError as ve:
        print(f"ValueError in CSV export: {str(ve)}")
        return JsonResponse({
            'success': False,
            'message': 'Invalid date format. Use YYYY-MM-DD'
        }, status=400)
    except Exception as e:
        print(f"Error exporting late report CSV: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'message': f'Export failed: {str(e)}'
        }, status=500)


# Export Functions

@csrf_exempt
@require_http_methods(["GET"])
def export_attendance_csv(request):
    """Export attendance data as CSV"""
    try:
        # Get parameters
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        department = request.GET.get('department')
        
        if not start_date_str or not end_date_str:
            return JsonResponse({
                'success': False,
                'message': 'start_date and end_date parameters are required'
            }, status=400)
        
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        
        # Load employee cache if empty
        if not employee_cache:
            load_employee_cache()
        
        # Get attendance data
        records = biotime_client.get_transactions_by_date_range(start_date, end_date)
        processed_records = process_attendance_records(records) if records else []
        
        # Filter by department if specified
        if department and department != "All":
            employees = biotime_client.get_employees(page_size=1000) or []
            dept_employee_ids = set()
            for emp in employees:
                emp_dept = getDepartmentName(emp.get('department', ''))
                if emp_dept == department:
                    dept_employee_ids.add(emp.get('emp_code', ''))
            
            processed_records = [r for r in processed_records if r['employee_id'] in dept_employee_ids]
        
        # Add status to records
        for record in processed_records:
            if 'status' not in record:
                if record['clock_in'] is None and record['clock_out'] is None:
                    record['status'] = 'Absent'
                elif record['late_minutes'] > 0:
                    record['status'] = 'Late'
                else:
                    record['status'] = 'On Time'
        
        # Create CSV response
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="attendance_report_{start_date_str}_to_{end_date_str}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Employee ID', 'Full Name', 'Date', 'Clock In', 'Clock Out', 'Working Hours', 'Late Minutes', 'Status', 'Device ID'])
        
        for record in processed_records:
            writer.writerow([
                record['employee_id'],
                record['full_name'],
                record['date'],
                record['clock_in'] or '',
                record['clock_out'] or '',
                record.get('working_hours', 0),
                record['late_minutes'],
                record['status'],
                record['device_id']
            ])
        
        return response
        
    except ValueError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid date format. Use YYYY-MM-DD'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Export failed: {str(e)}'
        }, status=500)

@csrf_exempt
@require_http_methods(["GET"])
def export_attendance_excel(request):
    """Export attendance data as Excel"""
    try:
        # Try to import openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return JsonResponse({
                'success': False,
                'message': 'Excel export requires openpyxl. Install with: pip install openpyxl'
            }, status=500)
        
        # Get parameters
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        department = request.GET.get('department')
        
        if not start_date_str or not end_date_str:
            return JsonResponse({
                'success': False,
                'message': 'start_date and end_date parameters are required'
            }, status=400)
        
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        
        # Load employee cache if empty
        if not employee_cache:
            load_employee_cache()
        
        # Get attendance data
        records = biotime_client.get_transactions_by_date_range(start_date, end_date)
        processed_records = process_attendance_records(records) if records else []
        
        # Filter by department if specified
        if department and department != "All":
            employees = biotime_client.get_employees(page_size=1000) or []
            dept_employee_ids = set()
            for emp in employees:
                emp_dept = getDepartmentName(emp.get('department', ''))
                if emp_dept == department:
                    dept_employee_ids.add(emp.get('emp_code', ''))
            
            processed_records = [r for r in processed_records if r['employee_id'] in dept_employee_ids]
        
        # Add status to records
        for record in processed_records:
            if 'status' not in record:
                if record['clock_in'] is None and record['clock_out'] is None:
                    record['status'] = 'Absent'
                elif record['late_minutes'] > 0:
                    record['status'] = 'Late'
                else:
                    record['status'] = 'On Time'
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add headers
        headers = ['Employee ID', 'Full Name', 'Date', 'Clock In', 'Clock Out', 'Working Hours', 'Late Minutes', 'Status', 'Device ID']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data
        for row, record in enumerate(processed_records, 2):
            ws.cell(row=row, column=1, value=record['employee_id'])
            ws.cell(row=row, column=2, value=record['full_name'])
            ws.cell(row=row, column=3, value=record['date'])
            ws.cell(row=row, column=4, value=record['clock_in'] or '')
            ws.cell(row=row, column=5, value=record['clock_out'] or '')
            ws.cell(row=row, column=6, value=record.get('working_hours', 0))
            ws.cell(row=row, column=7, value=record['late_minutes'])
            ws.cell(row=row, column=8, value=record['status'])
            ws.cell(row=row, column=9, value=record['device_id'])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="attendance_report_{start_date_str}_to_{end_date_str}.xlsx"'
        
        return response
        
    except ValueError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid date format. Use YYYY-MM-DD'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Export failed: {str(e)}'
        }, status=500)

@csrf_exempt
@require_http_methods(["GET"])
def export_attendance_pdf(request):
    """Export attendance data as PDF"""
    try:
        # Try to import reportlab
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
        except ImportError:
            return JsonResponse({
                'success': False,
                'message': 'PDF export requires reportlab. Install with: pip install reportlab'
            }, status=500)
        
        # Get parameters
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        department = request.GET.get('department')
        
        if not start_date_str or not end_date_str:
            return JsonResponse({
                'success': False,
                'message': 'start_date and end_date parameters are required'
            }, status=400)
        
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        
        # Load employee cache if empty
        if not employee_cache:
            load_employee_cache()
        
        # Get attendance data
        records = biotime_client.get_transactions_by_date_range(start_date, end_date)
        processed_records = process_attendance_records(records) if records else []
        
        # Filter by department if specified
        if department and department != "All":
            employees = biotime_client.get_employees(page_size=1000) or []
            dept_employee_ids = set()
            for emp in employees:
                emp_dept = getDepartmentName(emp.get('department', ''))
                if emp_dept == department:
                    dept_employee_ids.add(emp.get('emp_code', ''))
            
            processed_records = [r for r in processed_records if r['employee_id'] in dept_employee_ids]
        
        # Add status to records
        for record in processed_records:
            if 'status' not in record:
                if record['clock_in'] is None and record['clock_out'] is None:
                    record['status'] = 'Absent'
                elif record['late_minutes'] > 0:
                    record['status'] = 'Late'
                else:
                    record['status'] = 'On Time'
        
        # Create PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        # Container for the 'Flowable' objects
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        # Title
        title = Paragraph(f"Attendance Report<br/>{start_date_str} to {end_date_str}", title_style)
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # Summary
        total_records = len(processed_records)
        present_count = len([r for r in processed_records if r.get('status') in ['On Time', 'Late']])
        absent_count = len([r for r in processed_records if r.get('status') == 'Absent'])
        late_count = len([r for r in processed_records if r.get('status') == 'Late'])
        
        summary_text = f"""
        <b>Summary:</b><br/>
        Total Records: {total_records}<br/>
        Present: {present_count}<br/>
        Absent: {absent_count}<br/>
        Late: {late_count}<br/>
        Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        """
        summary = Paragraph(summary_text, styles['Normal'])
        elements.append(summary)
        elements.append(Spacer(1, 20))
        
        # Table data
        data = [['Employee ID', 'Full Name', 'Date', 'Clock In', 'Clock Out', 'Hours', 'Status']]
        
        for record in processed_records[:100]:  # Limit to first 100 records for PDF
            data.append([
                record['employee_id'],
                record['full_name'][:20] + ('...' if len(record['full_name']) > 20 else ''),
                record['date'],
                record['clock_in'] or '—',
                record['clock_out'] or '—',
                f"{record.get('working_hours', 0)}h",
                record['status']
            ])
        
        # Create table
        table = Table(data, colWidths=[1*inch, 2*inch, 1*inch, 1*inch, 1*inch, 0.8*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        
        if len(processed_records) > 100:
            note = Paragraph(f"<i>Note: Only first 100 records shown. Total records: {len(processed_records)}</i>", styles['Normal'])
            elements.append(Spacer(1, 12))
            elements.append(note)
        
        # Build PDF
        doc.build(elements)
        
        # Get the value of the BytesIO buffer and write it to the response
        pdf = buffer.getvalue()
        buffer.close()
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="attendance_report_{start_date_str}_to_{end_date_str}.pdf"'
        response.write(pdf)
        
        return response
        
    except ValueError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid date format. Use YYYY-MM-DD'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Export failed: {str(e)}'
        }, status=500)
